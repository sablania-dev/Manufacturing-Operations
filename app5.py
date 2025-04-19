# PART 1 - SETUP + DATA GENERATION

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import os
from io import BytesIO
from statsmodels.tsa.holtwinters import ExponentialSmoothing as HWES
import matplotlib.pyplot as plt
from openpyxl import Workbook  # using openpyxl instead of xlsxwriter for Excel downloads

st.set_page_config(page_title="GPU Manufacturing Operations Dashboard", layout="wide")

# --- Data Generation (Updated realistic values) ---
np.random.seed(42)

components = [
    'GPU Chip', 'Memory Module (GDDR6)', 'Memory Stack (HBM)',
    'PCB Board', 'VRM Controller', 'Power Stage',
    'Inductor', 'Capacitor', 'Cooling Fan', 'Heatsink Assembly'
]

inventory_data = pd.DataFrame({
    'Component': components,
    'Initial Inventory': np.random.randint(100, 800, len(components)),
    'Lead Time (weeks)': np.clip(np.random.randint(2, 8, len(components)), 1, 7),
    'Unit Cost ($)': [2500, 100, 2000, 15, 2, 1, 0.3, 0.2, 5, 10],
    'Ordering Cost ($)': [5000, 500, 2000, 1000, 300, 200, 100, 100, 500, 700],
    'Holding Cost %': [0.08, 0.1, 0.08, 0.05, 0.06, 0.06, 0.05, 0.05, 0.033, 0.033],
    'MOQ': [500, 2000, 300, 500, 1000, 2000, 4000, 4000, 500, 300]
})

inventory_data['Backorder Cost ($)'] = [
    150, 40, 200, 1, 150, 50, 40, 25, 35, 40
]  # comparable to holding cost, properly scaled

discount_types = np.random.choice(['AUD', 'ID', 'None'], size=len(components), p=[0.5, 0.3, 0.2])
inventory_data['Discount Type'] = discount_types
inventory_data['Discount Threshold (Units)'] = np.random.randint(300, 700, len(components))
inventory_data['Discount (%)'] = np.round(np.random.uniform(0.05, 0.15, len(components)) * 100, 1)

# Ensure 'EOQ (units)' column exists in inventory_data
if 'EOQ (units)' not in inventory_data.columns:
    eoq_results = []
    for idx, row in inventory_data.iterrows():
        D = 10000  # Example annual demand, replace with actual demand
        S = row['Ordering Cost ($)']
        B = row['Backorder Cost ($)']
        P_base = row['Unit Cost ($)']
        H_base = P_base * row['Holding Cost %']

        Q_star = np.sqrt((2 * D * S) * (P_base + B) / (H_base * B))
        eoq_results.append(Q_star)

    inventory_data['EOQ (units)'] = np.round(eoq_results)

# --- Demand Data: Monthly ---
months = pd.date_range(start='2025-01-01', periods=24, freq='M')
gaming_base_demand = np.random.poisson(lam=10000, size=24)
gaming_seasonality = np.sin(np.linspace(0, 6*np.pi, 24)) * 2000
gaming_growth = np.linspace(0, 2000, 24)
gaming_demand = np.maximum(gaming_base_demand + gaming_seasonality + gaming_growth, 0).astype(int)

dc_base_demand = np.random.poisson(lam=1000, size=24)
dc_growth = np.linspace(0, 800, 24)
dc_demand = np.maximum(dc_base_demand + dc_growth, 0).astype(int)

demand_df = pd.DataFrame({
    'Month': months,
    'Gaming_GPU_Demand': gaming_demand,
    'DataCenter_GPU_Demand': dc_demand
})
# PART 2 - DEMAND FORECASTING (Monthly Forecast and Download)

tab = st.sidebar.radio("Select Section", ["Demand Forecasting", "Inventory Management", "Material Requirement Planning"])

# ======================= DEMAND FORECASTING =========================
if tab == "Demand Forecasting":
    st.header("Realistic Demand Forecasting (Monthly Basis)")
    st.subheader("Simulated Monthly Demand Data")
    st.dataframe(demand_df)

    method = st.selectbox("Choose Forecasting Method", ["Moving Average", "Exponential Smoothing", "Holt-Winters"])

    if method == "Moving Average":
        window = st.slider("Select Moving Average Window", 2, 6, 3)
        demand_df['Gaming_GPU_Forecast'] = demand_df['Gaming_GPU_Demand'].rolling(window=window).mean().shift(1)
        demand_df['DataCenter_GPU_Forecast'] = demand_df['DataCenter_GPU_Demand'].rolling(window=window).mean().shift(1)
    elif method == "Exponential Smoothing":
        alpha = st.slider("Select Smoothing Constant Alpha", 0.1, 1.0, 0.5)
        demand_df['Gaming_GPU_Forecast'] = HWES(demand_df['Gaming_GPU_Demand']).fit(smoothing_level=alpha, optimized=False).fittedvalues
        demand_df['DataCenter_GPU_Forecast'] = HWES(demand_df['DataCenter_GPU_Demand']).fit(smoothing_level=alpha, optimized=False).fittedvalues
    else:
        gaming_model = HWES(demand_df['Gaming_GPU_Demand'], seasonal_periods=12, trend='add', seasonal='add').fit()
        demand_df['Gaming_GPU_Forecast'] = gaming_model.fittedvalues
        dc_model = HWES(demand_df['DataCenter_GPU_Demand'], seasonal_periods=12, trend='add', seasonal='add').fit()
        demand_df['DataCenter_GPU_Forecast'] = dc_model.fittedvalues

    # --- Save Button for Monthly Forecast ---
    def convert_df_to_csv(df):
        return df.to_csv(index=False).encode('utf-8')

    csv_forecast = convert_df_to_csv(demand_df[['Month', 'Gaming_GPU_Forecast', 'DataCenter_GPU_Forecast']])

    st.download_button(
        label="📥 Download Monthly Forecast CSV",
        data=csv_forecast,
        file_name='monthly_forecast.csv',
        mime='text/csv'
    )

    # --- Plotting Monthly Forecast ---
    st.subheader("Monthly Demand Forecast Plot")
    fig = px.line(
        demand_df,
        x='Month',
        y=['Gaming_GPU_Forecast', 'DataCenter_GPU_Forecast'],
        labels={'value': 'Units', 'Month': 'Month'},
        title="Monthly GPU Demand Forecast"
    )
    st.plotly_chart(fig, use_container_width=True)

    # ======================= INVENTORY MANAGEMENT =========================
elif tab == "Inventory Management":
    st.header("Inventory Management (EOQ with Backorders)")

    st.subheader("Step 1: Upload Saved Monthly Forecast")
    uploaded_forecast = st.file_uploader("Upload Monthly Forecast CSV", type=['csv'], key="forecast_inventory")
    
    if uploaded_forecast is not None:
        forecast_df = pd.read_csv(uploaded_forecast)
        st.success("Monthly Forecast loaded successfully!")
        st.dataframe(forecast_df)

        # Calculate Total Annual Demand (Sum of all forecasted months)
        total_annual_demand = forecast_df['Gaming_GPU_Forecast'].sum() + forecast_df['DataCenter_GPU_Forecast'].sum()
        st.write(f"**Total Annual Forecasted Demand for EOQ Calculations:** {int(total_annual_demand)} units")

        st.subheader("Step 2: EOQ Calculations (with Backorders and Discounts)")
        st.latex(r'Q^* = \sqrt{\frac{2DS}{H} \times \frac{P+B}{B}}')

        eoq_results = []
        total_costs = []
        effective_unit_costs = []

        for idx, row in inventory_data.iterrows():
            D = total_annual_demand
            S = row['Ordering Cost ($)']
            B = row['Backorder Cost ($)']
            P_base = row['Unit Cost ($)']
            H_base = P_base * row['Holding Cost %']

            discount_type = row['Discount Type']
            threshold = row['Discount Threshold (Units)']
            discount_percent = row['Discount (%)'] / 100

            # Estimate EOQ
            Q_star_estimate = np.sqrt((2 * D * S) * (P_base + B) / (H_base * B))

            if discount_type == 'AUD' and Q_star_estimate >= threshold:
                P_effective = P_base * (1 - discount_percent)
            elif discount_type == 'ID' and Q_star_estimate >= threshold:
                normal_units = threshold
                discounted_units = Q_star_estimate - threshold
                P_effective = (normal_units * P_base + discounted_units * P_base * (1 - discount_percent)) / Q_star_estimate
            else:
                P_effective = P_base

            H = P_effective * row['Holding Cost %']

            Q_star = np.sqrt((2 * D * S) * (P_effective + B) / (H * B))
            eoq_results.append(Q_star)
            effective_unit_costs.append(P_effective)

            ordering_cost = (D / Q_star) * S
            holding_cost = (Q_star / 2) * H
            backorder_cost = (D * (B / (P_effective + B))) * (H / 2)
            total_annual_cost = ordering_cost + holding_cost + backorder_cost+D * P_effective
            total_annual_cost = np.round(total_annual_cost, 2)  
            total_costs.append(total_annual_cost)
        print(inventory_data.head())

        inventory_data['Effective Unit Cost ($)'] = np.round(effective_unit_costs, 2)
        inventory_data['EOQ (units)'] = np.round(eoq_results)
        inventory_data['Total Annual Cost ($)'] = np.round(total_costs, 2)

        # --- Reset Backorder Cost to 5% of Unit Cost ---


        st.subheader("EOQ Results Table")
        st.dataframe(inventory_data[['Component', 'EOQ (units)', 'Effective Unit Cost ($)', 'Total Annual Cost ($)']])

        # --- Insert Detailed EOQ Analysis Table Here ---
        st.subheader("Step 4: Detailed EOQ Analysis Per Item")

        # Calculate missing fields
        detailed_table = inventory_data.copy()

        # 1. Calculate Annual Forecasted Demand per component
        bom_mapping = {
            'GPU Chip': (1, 1),
            'Memory Module (GDDR6)': (8, 0),
            'Memory Stack (HBM)': (0, 6),
            'PCB Board': (1, 1),
            'VRM Controller': (1, 2),
            'Power Stage': (8, 12),
            'Inductor': (8, 12),
            'Capacitor': (20, 40),
            'Cooling Fan': (2, 0),
            'Heatsink Assembly': (1, 1)
        }

        forecast_gaming = forecast_df['Gaming_GPU_Forecast'].sum()
        forecast_dc = forecast_df['DataCenter_GPU_Forecast'].sum()

        total_demands = []
        for comp in detailed_table['Component']:
            gaming_qty, dc_qty = bom_mapping.get(comp, (0, 0))
            total_qty = gaming_qty * forecast_gaming + dc_qty * forecast_dc
            total_demands.append(int(total_qty))

        detailed_table['Annual Forecasted Demand (units)'] = total_demands

        # 2. Holding Cost Per Unit
        detailed_table['Holding Cost ($/unit)'] = (detailed_table['Unit Cost ($)'] *detailed_table[ 'Holding Cost %']).round(2)

        # 3. Number of Orders Per Year
        detailed_table['Number of Orders Per Year'] = (detailed_table['Annual Forecasted Demand (units)'] / detailed_table['EOQ (units)']).round(2)

        # 4. Select Relevant Columns
        display_table = detailed_table[[
            'Component', 'Annual Forecasted Demand (units)', 'EOQ (units)',
            'Backorder Cost ($)', 'Holding Cost ($/unit)', 'Total Annual Cost ($)', 'Number of Orders Per Year'
        ]]

        st.dataframe(display_table)
# --- End of Detailed EOQ Analysis ---

        # --- Save EOQ Data ---
        csv_eoq = inventory_data[['Component', 'EOQ (units)', 'Effective Unit Cost ($)', 'Total Annual Cost ($)']].to_csv(index=False).encode('utf-8')
        # --- Reset Backorder Cost to 5% of Unit Cost ---
        

        st.download_button(
            label="📥 Download EOQ Data CSV",
            data=csv_eoq,
            file_name='eoq_data.csv',
            mime='text/csv'
        )

        # --- Step 3: Simulate New Q Values ---
        st.subheader("Step 3: Simulate Order Quantity and See New Total Cost")
        simulate_component = st.selectbox("Select Component to Simulate", inventory_data['Component'])
        simulate_row = inventory_data[inventory_data['Component'] == simulate_component].iloc[0]

        sim_Q = st.slider(f"Select Order Quantity (units) for {simulate_component}", 50, 5000, int(simulate_row['EOQ (units)']))

        # Recalculate Costs
        D = total_annual_demand
        S = simulate_row['Ordering Cost ($)']
        B = simulate_row['Backorder Cost ($)']
        P = simulate_row['Effective Unit Cost ($)']
        H = P * simulate_row['Holding Cost %']

        ordering_cost_sim = (D / sim_Q) * S
        holding_cost_sim = (sim_Q / 2) * H
        backorder_cost_sim = (D * (B / (P + B))) * (H / 2)
        total_simulated_cost = ordering_cost_sim + holding_cost_sim + backorder_cost_sim

        # Display
        original_total_cost = simulate_row['Total Annual Cost ($)']
        pct_change = ((total_simulated_cost - original_total_cost) / original_total_cost) * 100

        st.metric(label="Original Total Annual Cost ($)", value=f"${original_total_cost:,.2f}")
        st.metric(label="Simulated Total Annual Cost ($)", value=f"**${total_simulated_cost:,.2f}**")
        st.metric(label="Percentage Change", value=f"{pct_change:.2f}%")

        # --- Step 4: Plot Cost vs Order Quantity ---
        st.subheader("Cost vs Order Quantity Analysis")

        order_quantities = np.arange(50, 5000, 50)
        total_costs_list = []

        for Q in order_quantities:
            ordering_cost = (D / Q) * S
            holding_cost = (Q / 2) * H
            backorder_cost = (D * (B / (P + B))) * (H / 2)
            total_cost = ordering_cost + holding_cost + backorder_cost
            total_costs_list.append(total_cost)

        fig = px.line(
            x=order_quantities,
            y=total_costs_list,
            labels={'x': 'Order Quantity', 'y': 'Total Annual Cost ($)'},
            title=f"Total Cost vs Order Quantity for {simulate_component}"
        )
        st.plotly_chart(fig, use_container_width=True)
# ======================= MATERIAL REQUIREMENT PLANNING =========================
elif tab == "Material Requirement Planning":
    # --- Step 1: Upload Files ---
    st.subheader("Step 1: Upload Monthly Forecast and EOQ Data for MRP")
    uploaded_forecast = st.file_uploader("Upload Monthly Forecast CSV", type=['csv'], key="forecast_mrp2")
    uploaded_eoq = st.file_uploader("Upload EOQ Data CSV", type=['csv'], key="eoq_mrp2")

    if uploaded_forecast is not None and uploaded_eoq is not None:
        forecast_df = pd.read_csv(uploaded_forecast)
        eoq_df = pd.read_csv(uploaded_eoq)

        st.success("Monthly Forecast and EOQ Data loaded successfully!")

        # --- Step 2: Weekly Forecast Disaggregation ---
        week_numbers = []
        weekly_demand = []

        for idx, row in forecast_df.iterrows():
            month_total = row['Gaming_GPU_Forecast'] + row['DataCenter_GPU_Forecast']
            splits = np.random.dirichlet(np.ones(4), size=1).flatten()
            weekly_splits = (month_total * splits).round(0)

            for val in weekly_splits:
                week_numbers.append(f"Week {len(week_numbers)+1}")
                weekly_demand.append(val)

        # Only 52 weeks
        week_numbers = week_numbers[:52]
        weekly_demand = weekly_demand[:52]

        weekly_forecast = pd.DataFrame({
            'Week': week_numbers,
            'Weekly Total Demand': weekly_demand
        })

        st.subheader("Weekly Forecast (Total GPUs)")
        st.dataframe(weekly_forecast)

            # --- Step 3: Generate MRP Tables for EOQ and L4L ---
        st.subheader("Step 2: Generate MRP Tables for Each Component (EOQ and Lot-for-Lot)")

        # BOM Definition
        bom_df = pd.DataFrame({
            'Component': components,
            'Gaming_GPU_Quantity_per_Unit': [1, 8, 0, 1, 1, 8, 8, 20, 2, 1],
            'DataCenter_GPU_Quantity_per_Unit': [1, 0, 6, 1, 2, 12, 12, 40, 0, 1]
        })

        # Merge Initial Inventory and Lead Times
        bom_df = bom_df.merge(
            inventory_data[['Component', 'Initial Inventory', 'Lead Time (weeks)']],
            on='Component', how='left'
        )

        # Merge EOQ Values
        bom_df = bom_df.merge(
            eoq_df[['Component', 'EOQ (units)']],
            on='Component', how='left'
        )

        # Prepare dicts for final Excel output
        eoq_mrp_outputs = {}
        l4l_mrp_outputs = {}

        selected_components = ['PCB Board', 'VRM Controller', 'Cooling Fan']


        for idx, row in bom_df.iterrows():

            
            component = row['Component']


    # Skip components not selected
            if component not in selected_components:
                continue
            initial_inventory = row['Initial Inventory']
            lead_time_weeks = int(row['Lead Time (weeks)'])
            eoq_qty = int(row['EOQ (units)'])
            gamming_qty = row['Gaming_GPU_Quantity_per_Unit']
            dc_qty = row['DataCenter_GPU_Quantity_per_Unit']

            # Calculate Weekly Gross Requirements
            total_units_per_gpu = gamming_qty + dc_qty
            weekly_gross_req = (weekly_forecast['Weekly Total Demand'] * total_units_per_gpu).round(0).astype(int)

            # --- EOQ Policy MRP Generation ---
            gross = weekly_gross_req.tolist()
            scheduled = [0] * 52  # Assuming no scheduled receipts
            projected = []
            net_requirements = []
            planned_receipts = [0] * 52
            planned_releases = [0] * 52

            on_hand = initial_inventory
            for week in range(52):
                gross_req = gross[week]
                scheduled_rcpt = scheduled[week]

                if week == 0:
                    projected_inventory = on_hand + scheduled_rcpt - gross_req
                else:
                    projected_inventory = projected[week-1] + scheduled_rcpt - gross_req

                if projected_inventory >= 0:
                    net_req = 0
                else:
                    net_req = abs(projected_inventory)

                # Plan receipt if needed
                if net_req > 0:
                    order_qty = eoq_qty
                    planned_receipts[week] = order_qty
                    projected_inventory += order_qty

                projected.append(projected_inventory)
                net_requirements.append(net_req)

            # Now adjust Planned Order Releases (shifted by lead time)
            for week in range(52):
                release_week = week - lead_time_weeks
                if release_week >= 0:
                    planned_releases[release_week] = planned_receipts[week]

            eoq_mrp_outputs[component] = pd.DataFrame({
                'Week': [f"W{w+1}" for w in range(52)],
                'Gross Requirements': gross,
                'Scheduled Receipts': scheduled,
                'Projected on Hand': projected,
                'Net Requirements': net_requirements,
                'Planned Order Receipts': planned_receipts,
                'Planned Order Releases': planned_releases
            })

            # --- Lot-for-Lot (L4L) Policy MRP Generation ---
            gross = weekly_gross_req.tolist()
            scheduled = [0] * 52
            projected = []
            net_requirements = []
            planned_receipts = []
            planned_releases = []

            on_hand = initial_inventory
            for week in range(52):
                gross_req = gross[week]
                scheduled_rcpt = scheduled[week]

                if week == 0:
                    projected_inventory = on_hand + scheduled_rcpt - gross_req
                else:
                    projected_inventory = projected[week-1] + scheduled_rcpt - gross_req

                if projected_inventory >= 0:
                    net_req = 0
                else:
                    net_req = abs(projected_inventory)

                # Plan receipt exactly as net requirement (Lot-for-Lot)
                if net_req > 0:
                    order_qty = net_req
                    planned_receipts.append(order_qty)
                    projected_inventory += order_qty
                else:
                    planned_receipts.append(0)

                projected.append(projected_inventory)
                net_requirements.append(net_req)

            # Now adjust Planned Order Releases (shifted by lead time)
            l4l_planned_releases = [0] * 52
            for week in range(52):
                release_week = week - lead_time_weeks
                if release_week >= 0:
                    l4l_planned_releases[release_week] = planned_receipts[week]

            l4l_mrp_outputs[component] = pd.DataFrame({
                'Week': [f"W{w+1}" for w in range(52)],
                'Gross Requirements': gross,
                'Scheduled Receipts': scheduled,
                'Projected on Hand': projected,
                'Net Requirements': net_requirements,
                'Planned Order Receipts': planned_receipts,
                'Planned Order Releases': l4l_planned_releases
            })

            # --- Step 4: Download MRP Files for EOQ and L4L ---
    st.subheader("Step 3: Download Final MRP Excel Files")

    from openpyxl import Workbook
    #from openpyxl.writer.excel import save_virtual_workbook

    # Function to create Excel Bytes
    def create_excel_from_dict(mrp_dict, title):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        for comp_name, df in mrp_dict.items():
            ws = wb.create_sheet(title=comp_name[:31])
            for r_idx, row in enumerate(df.itertuples(index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx+1, column=c_idx, value=value)
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
        # --- THIS IS THE NEW CORRECT WAY ---
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # Download EOQ MRP
    eoq_excel = create_excel_from_dict(eoq_mrp_outputs, title="EOQ_MRP")
    st.download_button(
        label="📥 Download EOQ Policy MRP Excel",
        data=eoq_excel,
        file_name="MRP_EOQ.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Download Lot-for-Lot MRP
    l4l_excel = create_excel_from_dict(l4l_mrp_outputs, title="L4L_MRP")
    st.download_button(
        label="📥 Download Lot-for-Lot Policy MRP Excel",
        data=l4l_excel,
        file_name="MRP_LotForLot.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
        # --- Step 4: Download MRP Files (EOQ and L4L) ---
    

    # 📌 --- INSERT HERE the new Total MRP Cost Calculation ---
        # --- Step 5: Calculate Full Total Cost from MRP (EOQ Only) ---
        # --- Step 5: Correct Total MRP Cost Calculation based on Your Formula ---
    st.subheader("Step 4: Final Total MRP Cost (Correct Formula) (EOQ Policy Only)")

    holding_cost_map = (inventory_data.set_index('Component')['Unit Cost ($)'] * inventory_data.set_index('Component')['Holding Cost %'] / 52).round(4).to_dict()
    unit_cost_map = inventory_data.set_index('Component')['Unit Cost ($)'].to_dict()
    ordering_cost_map = inventory_data.set_index('Component')['Ordering Cost ($)'].to_dict()

    final_costs = []

    for comp_name in eoq_mrp_outputs.keys():
        df_eoq = eoq_mrp_outputs[comp_name]

        # Values needed
        unit_cost = unit_cost_map.get(comp_name, 0)
        ordering_cost = ordering_cost_map.get(comp_name, 0)
        weekly_holding_cost = holding_cost_map.get(comp_name, 0)

        # Calculate terms:
        total_orders = (df_eoq['Planned Order Releases'] > 0).sum()
        total_release_units = df_eoq['Planned Order Releases'].sum()
        avg_on_hand = df_eoq['Projected on Hand'].mean()

        # Apply your formula:
        total_ordering_cost = total_orders * ordering_cost
        total_purchase_cost = total_release_units * unit_cost
        total_holding_cost = weekly_holding_cost * avg_on_hand * 52

        total_mrp_cost = total_ordering_cost + total_purchase_cost + total_holding_cost

        final_costs.append({
            'Component': comp_name,
            'Ordering Cost ($)': round(total_ordering_cost, 2),
            'Purchase Cost ($)': round(total_purchase_cost, 2),
            'Holding Cost ($)': round(total_holding_cost, 2),
            'Total MRP Cost ($)': round(total_mrp_cost, 2)
        })

    final_mrp_costs_df = pd.DataFrame(final_costs)

    st.subheader("Final Detailed MRP Costs Per Component (Using Correct Formula)")
    st.dataframe(final_mrp_costs_df)

    grand_total_mrp_cost = final_mrp_costs_df['Total MRP Cost ($)'].sum()

    st.metric(label="🔵 GRAND TOTAL MRP COST (EOQ POLICY)", value=f"**${grand_total_mrp_cost:,.2f}**")